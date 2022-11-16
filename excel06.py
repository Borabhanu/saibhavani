from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'InvoiceNo'
ws['B1'] = 'StockCode'
ws['C1'] = 'Description'
ws['D1'] = 'Quantity'
ws['E1'] = 'InvoiceDate'
ws['F1'] = 'UnitPrice'
ws['G1'] = 'CustomerID'
ws['H1'] = 'Contry'
ws.append(['536365', '85123A', 'WHITE HANGING HEART' , '6' , '11-06-1999' , '2.55' , '17850' , 'United Kingdom'])
import datetime
ws['A2'] = datetime.datetime.now()
wb.save("sample2.xlsx")